from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "moph_focus_seed"

HEADERS = [
    "title","subtitle","summary",
    "mustKnow","examPoints","quickMemory","fullContent",
    "tags","coverEmoji","order","isPublished","publishedDate"
]

SEED = [
    {
        "title": "นโยบาย 30 บาทรักษาทุกที่",
        "subtitle": "ระบบหลักประกันสุขภาพถ้วนหน้า ยุคใหม่",
        "summary": "ประชาชนทุกคนสามารถรับบริการที่สถานพยาบาลใดก็ได้ทั่วประเทศ เพียงใช้บัตรประชาชนใบเดียว ไม่จำกัดสิทธิ์ตามภูมิลำเนาอีกต่อไป นโยบายสำคัญของรัฐบาลปี 2566",
        "mustKnow": "**หลักการ**: ใช้บัตรประชาชน 1 ใบ รับบริการได้ทุกสถานพยาบาลในระบบ\n**เริ่มต้น**: นำร่อง 4 จังหวัด ปี 2566 ได้แก่ แพร่ ร้อยเอ็ด เพชรบุรี นราธิวาส\n**ปัจจุบัน**: ขยายครอบคลุมทั่วประเทศ\n**สิทธิ์ที่ใช้**: บัตรทอง (UC), ประกันสังคม, ข้าราชการ",
        "examPoints": "- อาจถามว่า **นโยบายนี้เริ่มปีไหน** และจังหวัดนำร่องที่ไหน\n- ถามว่า **ความแตกต่าง** จากระบบบัตรทองเดิมคืออะไร\n- ถามเกี่ยวกับ **ระบบ NHSO** และการเชื่อมต่อฐานข้อมูล",
        "quickMemory": "30 บาท = **บัตรประชาชนใบเดียว ทุกที่ทั่วไทย**\nจำ: 4 จังหวัดนำร่อง = แพร่ ร้อยเอ็ด เพชรบุรี นราธิวาส",
        "fullContent": "## ที่มาของนโยบาย\nนโยบาย 30 บาทรักษาทุกที่ เป็นการพัฒนาต่อยอดจากโครงการหลักประกันสุขภาพถ้วนหน้า (บัตรทอง) ที่เริ่มต้นตั้งแต่ปี 2545\n\n## เทคโนโลยีที่รองรับ\n- ระบบ **NHSO Digital Platform**\n- การเชื่อมต่อฐานข้อมูล **43 แฟ้ม**\n- **Health ID** — เลขบัตรประชาชน = เลขสุขภาพ",
        "tags": "Quick Win|นโยบาย|ข่าวที่ควรรู้",
        "coverEmoji": "🏥",
        "order": 1,
        "isPublished": 1,
        "publishedDate": "2025-01-15",
    },
    {
        "title": "Digital Health และ AI ด้านสาธารณสุข",
        "subtitle": "การนำเทคโนโลยีดิจิทัลมาใช้ในระบบสุขภาพไทย",
        "summary": "ประเทศไทยกำลังเปลี่ยนผ่านสู่ระบบสุขภาพดิจิทัล ทั้ง AI วินิจฉัยโรค Telemedicine และ Health Data Platform ซึ่งเป็นประเด็นที่ออกสอบบ่อยมากในปีล่าสุด",
        "mustKnow": "**Digital Health คืออะไร**: การนำเทคโนโลยีดิจิทัลมาพัฒนาระบบสุขภาพ\n**AI ในสาธารณสุข**:\n- วินิจฉัยภาพถ่ายรังสี (X-ray, CT)\n- คัดกรอง Diabetic Retinopathy\n- ทำนายความเสี่ยงโรค\n**ระบบสำคัญของไทย**:\n- HosXP / HIS (Hospital Information System)\n- Health Data Center (HDC)\n- 43 แฟ้มข้อมูลสุขภาพ",
        "examPoints": "- ถามว่า **43 แฟ้ม** คืออะไร ส่งไปที่ไหน\n- **HDC** = Health Data Center = คลังข้อมูลสุขภาพกระทรวง\n- ถามเกี่ยวกับ **PDPA** กับข้อมูลสุขภาพ\n- ถามว่า AI ช่วย **คัดกรอง** โรคอะไรได้บ้าง",
        "quickMemory": "Digital Health = **HIS + HDC + AI + Telemedicine**\n43 แฟ้ม → ส่งให้ สปสช. ทุกเดือน",
        "fullContent": "## 43 แฟ้มข้อมูลสาธารณสุข\nเป็นมาตรฐานการส่งข้อมูลจากโรงพยาบาล/สถานีอนามัย ไปยัง สปสช. และกระทรวงสาธารณสุข\n\n## PDPA กับข้อมูลสุขภาพ\n- ข้อมูลสุขภาพ = **ข้อมูลอ่อนไหว (Sensitive Data)**\n- ต้องได้รับความยินยอมอย่างชัดแจ้ง\n- ผู้ป่วยมีสิทธิ์ขอดู แก้ไข และลบข้อมูล",
        "tags": "Digital Health|นโยบาย|Quick Win",
        "coverEmoji": "💻",
        "order": 2,
        "isPublished": 1,
        "publishedDate": "2025-02-01",
    },
    {
        "title": "วิกฤต PM2.5 และนโยบายอากาศสะอาด",
        "subtitle": "มลพิษทางอากาศกับผลกระทบต่อสุขภาพ",
        "summary": "PM2.5 เกินมาตรฐานในหลายจังหวัดของไทย โดยเฉพาะภาคเหนือช่วง ม.ค.-เม.ย. ส่งผลต่อระบบทางเดินหายใจและหัวใจ รัฐบาลออกนโยบายอากาศสะอาดและ พ.ร.บ.อากาศสะอาด",
        "mustKnow": "**ค่ามาตรฐาน PM2.5 ของไทย**:\n- ค่าเฉลี่ย 24 ชั่วโมง: ≤ 37.5 μg/m³ (มาตรฐานเดิม)\n- ปรับใหม่ปี 2566: ≤ 25 μg/m³ (ใกล้เคียง WHO)\n**ผลต่อสุขภาพ**:\n- ระบบทางเดินหายใจ: หอบหืด, ปอดอักเสบ\n- หัวใจและหลอดเลือด: หัวใจวาย, หลอดเลือดสมอง\n- กลุ่มเสี่ยง: เด็ก, ผู้สูงอายุ, ผู้ป่วยโรคเรื้อรัง",
        "examPoints": "- ถามค่ามาตรฐาน PM2.5 ของไทย vs WHO (WHO = 15 μg/m³)\n- กลุ่มเสี่ยงสูง = เด็ก, ผู้สูงอายุ, ผู้ป่วยหัวใจ/ปอด\n- มาตรการระยะสั้น vs ระยะยาวในการลด PM2.5\n- ความสัมพันธ์ PM2.5 กับ **โรคหัวใจและหลอดเลือด**",
        "quickMemory": "PM2.5 ไทย = **37.5** (เดิม) → **25** μg/m³ (ใหม่)\nWHO = **15** μg/m³\nจำ: ยิ่งเลขน้อย = มาตรฐานเข้มกว่า",
        "fullContent": "## แหล่งกำเนิด PM2.5 ในไทย\n- **ภาคเหนือ**: การเผาป่า เผาเศษพืช\n- **กรุงเทพฯ**: การจราจร โรงงานอุตสาหกรรม\n- **ภาคตะวันออก**: โรงงานอุตสาหกรรมปิโตรเคมี\n\n## มาตรการของรัฐ\n- ห้ามเผาในที่โล่ง (ช่วง 60 วันห้ามเผา)\n- แอป AirBKK, Air4Thai ติดตามค่าฝุ่น\n- N95 Mask แจกกลุ่มเสี่ยง",
        "tags": "PM2.5|โรคและภัยสุขภาพ|กฎหมาย",
        "coverEmoji": "🌫",
        "order": 3,
        "isPublished": 1,
        "publishedDate": "2025-02-15",
    },
    {
        "title": "NCD — โรคไม่ติดต่อเรื้อรัง",
        "subtitle": "เบาหวาน ความดัน หัวใจ มะเร็ง — ภัยเงียบที่ออกสอบทุกปี",
        "summary": "NCD (Non-Communicable Disease) คือกลุ่มโรคที่เป็นสาเหตุการตายอันดับ 1 ของไทย ได้แก่ โรคหัวใจ มะเร็ง เบาหวาน และความดันโลหิตสูง นโยบายป้องกันและควบคุม NCD เป็นประเด็นสอบสำคัญ",
        "mustKnow": "**NCD 4 โรคหลัก**:\n- โรคหัวใจและหลอดเลือด (CVD)\n- มะเร็ง\n- เบาหวาน (DM)\n- โรคปอดเรื้อรัง (COPD)\n\n**ปัจจัยเสี่ยงร่วม**:\n- บุหรี่, แอลกอฮอล์\n- อาหาร (เค็ม หวาน มัน)\n- ขาดการออกกำลังกาย\n\n**เป้าหมาย WHO**: ลด NCD ก่อนวัยอันควร 25% ภายในปี 2568",
        "examPoints": "- ถามว่า NCD ย่อมาจากอะไร และมีกี่กลุ่มโรค\n- **Shared Risk Factor** = ปัจจัยเสี่ยงร่วมกันของ NCD ทุกโรค\n- นโยบาย **3อ2ส** ของกระทรวงสาธารณสุข\n- เป้าหมาย **SDG 3.4** ที่เกี่ยวกับ NCD\n- อัตราตาย NCD ในไทย = 75% ของการตายทั้งหมด",
        "quickMemory": "NCD = **ไม่ติดต่อ ไม่หาย แต่ป้องกันได้**\n4 โรคหลัก = หัวใจ มะเร็ง เบาหวาน ปอด\n3อ2ส = อาหาร ออกกำลังกาย อารมณ์ + สูบบุหรี่ สุรา",
        "fullContent": "## สถานการณ์ NCD ในไทย\n- NCD เป็นสาเหตุการตาย **75%** ของการตายทั้งหมด\n- ค่าใช้จ่ายรักษา NCD = **~60%** ของค่าใช้จ่ายสุขภาพทั้งหมด\n\n## นโยบาย 3อ2ส\n- **3อ**: อาหาร (ลดหวานมันเค็ม), ออกกำลังกาย (150 นาที/สัปดาห์), อารมณ์\n- **2ส**: ไม่สูบบุหรี่, ไม่ดื่มสุรา\n\n## SDG 3.4\nลดการเสียชีวิตก่อนวัยอันควรจาก NCD ลง 1 ใน 3 ภายในปี 2573",
        "tags": "โรคและภัยสุขภาพ|นโยบาย|Quick Win",
        "coverEmoji": "🫀",
        "order": 4,
        "isPublished": 1,
        "publishedDate": "2025-03-01",
    },
    {
        "title": "Health Station และ Telemedicine",
        "subtitle": "เครื่องตรวจสุขภาพอัตโนมัติและการแพทย์ทางไกล",
        "summary": "Health Station เป็นเครื่องตรวจสุขภาพอัตโนมัติที่กระจายในชุมชน ส่วน Telemedicine คือการพบแพทย์ผ่านวิดีโอคอล ทั้งสองเป็นหัวใจของนโยบาย 30 บาทรักษาทุกที่ยุคดิจิทัล",
        "mustKnow": "**Health Station คืออะไร**:\n- เครื่องตรวจสุขภาพอัตโนมัติ วัด BP, น้ำหนัก, ออกซิเจน\n- เชื่อมต่อกับระบบ HIS และ PHR\n- ตั้งอยู่ในร้านยา สถานีอนามัย ชุมชน\n\n**Telemedicine**:\n- นิยาม: การให้บริการทางการแพทย์ผ่านสื่ออิเล็กทรอนิกส์\n- กฎหมาย: ประกาศแพทยสภา ปี 2563\n- ใช้ได้: ติดตามอาการโรคเรื้อรัง, ใบสั่งยา\n- ไม่ควรใช้: ผู้ป่วยฉุกเฉิน, ตรวจครั้งแรก",
        "examPoints": "- ถามนิยาม **Telemedicine** vs Telehealth\n- ข้อบ่งชี้และข้อห้ามของ Telemedicine\n- **Health Station** เชื่อมกับระบบอะไร\n- กฎหมายที่เกี่ยวข้อง: ประกาศแพทยสภา ปี 2563\n- บทบาทใน **30 บาทรักษาทุกที่**",
        "quickMemory": "Telemedicine = **พบแพทย์ได้ ไม่ต้องเดินทาง**\nใช้ได้: โรคเรื้อรังที่ควบคุมได้\nไม่ใช้: ฉุกเฉิน, ตรวจแรก",
        "fullContent": "## Telemedicine vs Telehealth\n- **Telemedicine**: แพทย์ให้การรักษา/สั่งยาผ่านระบบดิจิทัล\n- **Telehealth**: กว้างกว่า ครอบคลุมทั้งการศึกษา ป้องกันโรค ติดตามผล\n\n## Health Station ในอนาคต\n- เชื่อมต่อ AI วินิจฉัยเบื้องต้น\n- ส่งข้อมูลแบบ Real-time\n- ลดภาระ OPD โรงพยาบาล\n\n## ข้อกังวล\n- ความปลอดภัยของข้อมูล (PDPA)\n- ความเหลื่อมล้ำด้านดิจิทัล (Digital Divide)",
        "tags": "Digital Health|นโยบาย|ข่าวที่ควรรู้",
        "coverEmoji": "📱",
        "order": 5,
        "isPublished": 1,
        "publishedDate": "2025-03-15",
    },
]

header_fill = PatternFill("solid", start_color="0369A1", end_color="0369A1")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
border = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="DDDDDD"),
)

for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

alt_fill = PatternFill("solid", start_color="F0F9FF", end_color="F0F9FF")
data_font = Font(name="Arial", size=10)
wrap_align = Alignment(vertical="top", wrap_text=True)

for r_idx, row_data in enumerate(SEED, 2):
    row_fill = alt_fill if r_idx % 2 == 0 else None
    for c_idx, col in enumerate(HEADERS, 1):
        val = row_data.get(col, "")
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = data_font
        cell.alignment = wrap_align
        if row_fill:
            cell.fill = row_fill
        cell.border = border

col_widths = {
    "title": 30, "subtitle": 30, "summary": 40,
    "mustKnow": 50, "examPoints": 45, "quickMemory": 35, "fullContent": 55,
    "tags": 30, "coverEmoji": 12, "order": 8, "isPublished": 12, "publishedDate": 15,
}
for c_idx, col in enumerate(HEADERS, 1):
    ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col, 20)

ws.row_dimensions[1].height = 30
for r in range(2, len(SEED) + 2):
    ws.row_dimensions[r].height = 80

ws.freeze_panes = "A2"

out_path = r"C:\Users\UNS_CT\Desktop\online-exam\moph_focus_seed.xlsx"
wb.save(out_path)
print("Saved:", out_path)
